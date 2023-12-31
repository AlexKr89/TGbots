import openpyxl
from datetime import datetime

class Database:
    def __init__(self, filename):
        self.filename = filename

    def get_events(self):
        wb = openpyxl.load_workbook(self.filename)
        sheet = wb.active
        events = [(row[0], row[1], row[2]) for row in sheet.iter_rows(min_row=2, values_only=True) if row[0] and row[1] and row[2]]
        wb.close()
        return events

class RegistrationDatabase:
    def __init__(self, filename):
        self.filename = filename

    def save_registration(self, event_name, user_info, user_phone):
        wb = openpyxl.load_workbook(self.filename)
        sheet = wb.active

        registration_data = [user_info, user_phone, datetime.now()]
        sheet.append([event_name, *registration_data])

        wb.save(self.filename)
        wb.close()
